#!/usr/bin/env python3
import time
import argparse
from openpyxl import load_workbook
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.options import Options

def get_map_center(driver, nodeNo):
    """
    指定した nodeNo を入力して検索し、マップの中央座標を取得する。
    ※ 既に「mapBtnLayerWindow」と「ui-id-4」をクリック済みの状態で呼び出す前提です。
    """
    wait = WebDriverWait(driver, 10)
    
    # id="txt_node_no" の入力欄に nodeNo を入力
    input_node = wait.until(EC.presence_of_element_located((By.ID, "txt_node_no")))
    input_node.clear()
    input_node.send_keys(nodeNo)
    
    # id="btn_search_node_pos" の検索ボタンがクリック可能になるのを待つ
    btn_search = wait.until(EC.element_to_be_clickable((By.ID, "btn_search_node_pos")))
    btn_search.click()
    
    # マップの更新待ち（必要に応じて調整）
    time.sleep(500 / 1000)  # 2秒待機
    
    # JavaScript を実行して map.getCenter() の結果を取得
    center = driver.execute_script("return map.getCenter();")
    return center

def process_excel(file_path, start_arg, end_arg):
    """
    Excelファイルを読み込み、各行のA列（交差点番号）に対してマップ検索を行い、
    得られた座標（lat, lng）をそれぞれF列、G列に書き込む。
    開始行と終了行はコマンドライン引数で指定（負の値の場合は末尾からのオフセットとして処理）
    """
    # headless モードの Chrome オプションを設定
    chrome_options = Options()
    chrome_options.add_argument("--headless")
    chrome_options.add_argument("--disable-gpu")
    chrome_options.add_argument("--no-sandbox")
    
    driver = webdriver.Chrome(options=chrome_options)
    driver.get("https://lg47web.tokusya.ktr.mlit.go.jp/binrangis/main.php")
    wait = WebDriverWait(driver, 10)
    
    # 最初に必要なボタンをクリックしておく
    btn_map_layer = wait.until(EC.element_to_be_clickable((By.ID, "mapBtnLayerWindow")))
    btn_map_layer.click()
    time.sleep(500 / 1000)
    
    btn_ui_id4 = wait.until(EC.element_to_be_clickable((By.ID, "ui-id-4")))
    btn_ui_id4.click()
    time.sleep(500 / 1000)
    driver.execute_script("map.setView([41.75799552006108, 140.72010040283206], 19);")
    
    # Excel ファイルを読み込む
    wb = load_workbook(file_path)
    ws = wb.active  # アクティブなシートを利用
    
    # 最終行の取得
    last_row = ws.max_row
    
    # 引数で与えた start, end を実際の行番号に変換
    # データは1行目がヘッダーのため、最低2行目から処理
    def conv_index(idx):
        if idx < 0:
            return last_row + idx + 1  # 例：last_row=100, idx=-10 → 100 -10 +1 = 91
        else:
            return idx
    start_row = conv_index(start_arg)
    end_row = conv_index(end_arg)
    if start_row < 2:
        start_row = 2
    if end_row < start_row:
        print("終了行が開始行より前です。処理を終了します。")
        driver.quit()
        return

    print(f"処理対象の行: {start_row} 行目～{end_row} 行目 (全体: {last_row} 行)")

    # 各行について、指定範囲の行を処理
    counter = start_row
    for row in ws.iter_rows(min_row=start_row, max_row=end_row):
        node_cell = row[0]  # A列のセル（交差点番号）
        if node_cell.value is None:
            continue

        # F列とG列（6,7列）が既に記載されている場合はスキップ
        lat_cell = ws.cell(row=node_cell.row, column=6)
        lng_cell = ws.cell(row=node_cell.row, column=7)
        if lat_cell.value is not None and lng_cell.value is not None:
            print(f"{counter} 交差点番号 {node_cell.value} は既に処理済みのためスキップします。")
            counter += 1
            continue

        nodeNo = str(node_cell.value)
        try:
            center = get_map_center(driver, nodeNo)
            if center is None:
                print(f"交差点番号 {nodeNo} の座標が取得できませんでした。")
                continue
            lat = center.get("lat")
            lng = center.get("lng")
            ws.cell(row=node_cell.row, column=6, value=lat)
            ws.cell(row=node_cell.row, column=7, value=lng)
            print(f"{counter} 交差点番号 {nodeNo}: lat={lat}, lng={lng}")
        except Exception as e:
            print(f"{counter} 交差点番号 {nodeNo} の処理中にエラーが発生")
        counter += 1

    driver.quit()
    wb.save(file_path)
    print("Excelファイルに結果を保存しました。")

if __name__ == "__main__":
    parser = argparse.ArgumentParser(
        description="指定された開始行と終了行の範囲内で、Excelの交差点番号からマップ検索して座標を取得します。"
    )
    parser.add_argument("start", type=int, help="開始行（負の値の場合は末尾からのオフセット）")
    parser.add_argument("end", type=int, help="終了行（負の値の場合は末尾からのオフセット、終了行は含む）")
    parser.add_argument("--file", "-f", type=str, default="intersection.xlsx", help="対象Excelファイル（デフォルト: intersection.xlsx）")
    args = parser.parse_args()
    
    process_excel(args.file, args.start, args.end)
